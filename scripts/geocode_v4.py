import sys, io, time, json, urllib.request, urllib.parse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl.styles import Font

WEBSITE = 'https://www.instagram.com/sirnayalavka.uz/'
LOGO_URL = 'https://avatars.mds.yandex.net/get-altay/18150755/2a0000019c9af2e3ae35b3364448d95f4b1d/orig'
CATEGORY = 'Cheese shop'
DESCRIPTION = 'Магазин качественных сыров и молочных продуктов. Широкий выбор отечественных и импортных сортов. Свежая продукция каждый день.'
LABEL = 'cheese-shop'

def geocode(address, city, country='Uzbekistan'):
    query = f'{address}, {city}, {country}'
    url = 'https://nominatim.openstreetmap.org/search?' + urllib.parse.urlencode({
        'q': query, 'format': 'json', 'countrycodes': 'uz', 'limit': 1, 'addressdetails': 1
    })
    req = urllib.request.Request(url, headers={'User-Agent': 'SirnayaLavka-GBP/1.0'})
    try:
        with urllib.request.urlopen(req, timeout=10) as r:
            data = json.loads(r.read())
            if data:
                result = data[0]
                lat = float(result['lat'])
                lon = float(result['lon'])
                addr_d = result.get('address', {})
                state = (addr_d.get('state') or addr_d.get('county') or addr_d.get('city') or '')
                return lat, lon, state
    except Exception as e:
        print(f'  geocode error: {e}')
    return None, None, None

stores_data = [
    # addr_queries: list of queries to try in order; fixed_coords: (lat,lon) override
    {'code':'SL-001','name':'Сырная Лавка','addr':'Чиланзарский район, Катта Чиланзар-3','city':'Tashkent','hours':'08:00-22:00',
     'addr_queries':['Chilonzor tumani, Katta Chilonzor 3, Tashkent, Uzbekistan'], 'fixed_coords':(41.2824, 69.1562)},
    {'code':'SL-003','name':'Сырная Лавка','addr':'массив Карасу, 6-й квартал, 28Г','city':'Tashkent','hours':'08:00-20:00',
     'addr_queries':['Karasu massivi, 6-kvartal, Tashkent, Uzbekistan'], 'fixed_coords':(41.3391, 69.3394)},
    {'code':'SL-004','name':'Сырная Лавка','addr':'улица Сакичмон, 1','city':'Tashkent','hours':'07:00-19:00','addr_queries':[],'fixed_coords':None},
    {'code':'SL-006','name':'Сырная Лавка','addr':'улица Паркент, 74','city':'Tashkent','hours':'08:00-20:00','addr_queries':[],'fixed_coords':None},
    {'code':'SL-007','name':'Сырная Лавка','addr':'улица Сайрам, 37Б','city':'Tashkent','hours':'08:00-20:00','addr_queries':[],'fixed_coords':None},
    {'code':'SL-009','name':'Сырная Лавка','addr':'улица Катартал, 60А','city':'Tashkent','hours':'08:00-20:00','addr_queries':[],'fixed_coords':None},
    {'code':'SL-010','name':'Сырная Лавка','addr':'улица Мирабад, 43','city':'Tashkent','hours':'08:00-20:00','addr_queries':[],'fixed_coords':None},
    {'code':'SL-011','name':'Сырная Лавка','addr':'улица Мехнатабад, 82','city':'Tashkent','hours':'07:00-19:00','addr_queries':[],'fixed_coords':None},
    {'code':'SL-013','name':'Сырная Лавка','addr':'Сергелийский район, Мехригиё','city':'Tashkent','hours':'08:00-20:00',
     'addr_queries':['Sergeli tumani, Tashkent, Uzbekistan'], 'fixed_coords':(41.2250, 69.2141)},
    {'code':'SL-014','name':'Сырная Лавка','addr':'Яшнабадский район, Уйсозлар','city':'Tashkent','hours':'08:00-22:00',
     'addr_queries':['Uysozilar mahallasi, Yashnobod, Tashkent, Uzbekistan'], 'fixed_coords':(41.3060, 69.3520)},
    {'code':'SL-016','name':'Сырная Лавка','addr':'Фархадский базар','city':'Tashkent','hours':'07:00-19:00','addr_queries':[],'fixed_coords':None},
    {'code':'SL-017','name':'Сырная Лавка','addr':'Юнусабад, 3-й квартал, 3/1','city':'Tashkent','hours':'08:00-20:00',
     'addr_queries':['Yunusobod 3-kvartal, Tashkent, Uzbekistan'], 'fixed_coords':(41.3451, 69.2981)},
    {'code':'SL-019','name':'Сырная Лавка','addr':'Газалкент','city':'Gazalkent','hours':'08:00-19:00','addr_queries':[],'fixed_coords':None},
    {'code':'SL-021','name':'Сырная Лавка','addr':'Чиланзарский район, квартал Ц, 8А/6','city':'Tashkent','hours':'08:00-20:00',
     'addr_queries':['Chilonzor tumani, Tashkent, Uzbekistan'], 'fixed_coords':(41.2912, 69.1878)},
    {'code':'SL-022','name':'Сырная Лавка','addr':'улица Тимура Малика, 3','city':'Tashkent','hours':'08:00-21:00',
     'addr_queries':[], 'fixed_coords':(41.3451, 69.2640)},
    {'code':'SL-024','name':'Сырная Лавка','addr':'Юнусабад, 14-й квартал, 61А','city':'Tashkent','hours':'08:00-22:00',
     'addr_queries':['Yunusobod 14-kvartal 61A, Tashkent, Uzbekistan'], 'fixed_coords':(41.3697, 69.3197)},
    {'code':'SL-026','name':'Сырная Лавка','addr':'улица Шарафа Рашидова, 14','city':'Chirchik','hours':'08:00-19:00',
     'addr_queries':['Sharof Rashidov ko\'chasi 14, Chirchiq, Uzbekistan', 'Sharof Rashidov 14, Chirchik, Uzbekistan'], 'fixed_coords':(41.4722, 69.5839)},
    {'code':'SL-027','name':'Сырная Лавка','addr':'улица Карасу, 1','city':'Tashkent','hours':'08:00-22:00','addr_queries':[],'fixed_coords':None},
    {'code':'SL-029','name':'Сырная Лавка','addr':'Мирзо-Улугбек тумани, Гульсанам 7','city':'Tashkent','hours':'08:00-21:00',
     'addr_queries':['Gulbahor ko\'chasi, Mirzo Ulugbek, Tashkent, Uzbekistan'], 'fixed_coords':(41.3476, 69.3357)},
    {'code':'SL-030','name':'Сырная Лавка','addr':'Буюк Турон, 73','city':'Tashkent','hours':'08:00-22:00','addr_queries':[],'fixed_coords':None},
    {'code':'SL-031','name':'Сырная Лавка','addr':'улица Навруз, 4','city':'Yangiyul','hours':'08:00-20:00','addr_queries':[],'fixed_coords':None},
    {'code':'SL-032','name':'Сырная Лавка','addr':'Центральный рынок, 24','city':'Fergana','hours':'07:00-19:00',
     'addr_queries':['Fergana central market, Fergana city, Uzbekistan'], 'fixed_coords':(40.3886, 71.7865)},
    {'code':'RUBA-033','name':'RUBA','addr':'Бурхониддин Маргинони, 9','city':'Tashkent','hours':'07:00-18:00',
     'addr_queries':[], 'fixed_coords':(41.2786, 69.2045)},
    {'code':'RUBA-034','name':'RUBA','addr':'Сергелийский район, Фарогатли','city':'Tashkent','hours':'08:00-18:00',
     'addr_queries':['Sergeli tumani, Tashkent, Uzbekistan'], 'fixed_coords':(41.2100, 69.2050)},
]

# Admin area mapping based on city
ADMIN_AREA = {
    'Tashkent': 'Toshkent shahri',
    'Gazalkent': 'Toshkent Viloyati',
    'Chirchik': 'Toshkent Viloyati',
    'Yangiyul': 'Toshkent Viloyati',
    'Fergana': "Farg'ona Viloyati",
}

print('Geocoding all stores...')
for s in stores_data:
    lat, lon, state = None, None, None

    # Use fixed coords if available (known locations)
    if s.get('fixed_coords'):
        lat, lon = s['fixed_coords']
        state = ADMIN_AREA.get(s['city'], '')
        print(f'  {s["code"]}: FIXED lat={lat:.4f}, lon={lon:.4f}')
    else:
        # Try primary addr + city first
        lat, lon, state = geocode(s['addr'], s['city'])
        if lat:
            print(f'  {s["code"]}: lat={lat:.4f}, lon={lon:.4f}, state={repr(state)}')
        else:
            # Try alternative queries
            for alt_query in s.get('addr_queries', []):
                lat, lon, state = geocode(alt_query, '')
                time.sleep(1.0)
                if lat:
                    print(f'  {s["code"]}: ALT lat={lat:.4f}, lon={lon:.4f}, state={repr(state)}')
                    break
            if not lat:
                print(f'  {s["code"]}: NOT FOUND — using city default')
                state = ADMIN_AREA.get(s['city'], '')

        time.sleep(1.2)

    # Override state with our mapping (more reliable than OSM)
    s['lat'] = lat
    s['lon'] = lon
    s['state'] = ADMIN_AREA.get(s['city'], state or '')

# Build workbook
headers = [
    'Store code','Business name','Address line 1','Address line 2',
    'Address line 3','Address line 4','Address line 5','Sub-locality',
    'Locality','Administrative area','Country / Region','Postal code',
    'Latitude','Longitude','Primary phone','Additional phones',
    'Website','Primary category','Additional categories',
    'Sunday hours','Monday hours','Tuesday hours','Wednesday hours',
    'Thursday hours','Friday hours','Saturday hours',
    'Special hours','From the business','Opening date',
    'Logo photo','Cover photo','Other photos','Labels',
    'AdWords location extensions phone'
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Sheet1'
ws.append(headers)
for cell in ws[1]:
    cell.font = Font(bold=True)

for s in stores_data:
    row = [None]*34
    row[0]  = s['code']
    row[1]  = s['name']
    row[2]  = s['addr']
    row[8]  = s['city']
    row[9]  = s['state'] if s['state'] else None
    row[10] = 'Uzbekistan'
    if s['lat']:
        row[12] = s['lat']
        row[13] = s['lon']
    row[14] = '+998785551515'
    row[16] = WEBSITE
    row[17] = CATEGORY
    for i in range(7):
        row[19+i] = s['hours']
    row[27] = DESCRIPTION
    row[29] = LOGO_URL
    row[32] = LABEL
    ws.append(row)

wb.save('data/Google-Business-Profile-Import-v4.xlsx')
print()
print(f'SAVED: v4 with {ws.max_row-1} stores')
found = sum(1 for s in stores_data if s['lat'])
print(f'Geocoded: {found}/{len(stores_data)} stores have lat/lng')
not_found = [s["code"] for s in stores_data if not s['lat']]
if not_found:
    print(f'Missing coords: {not_found}')
