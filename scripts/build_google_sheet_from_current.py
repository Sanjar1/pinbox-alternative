import json
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


EXCEL_PATH = Path("data/stores_audit.xlsx")
ATTEMPTS_PATH = Path("data/google_claim_attempts_2026-02-27.json")


def solid_fill(argb: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=argb)


def color_font(argb: str) -> Font:
    return Font(color=argb)


def listing(name: str, address: str, status: str, fid: str, store_code: str, profile_url: str):
    return {
        "name": name,
        "address": address,
        "status": status,
        "fid": fid,
        "store_code": store_code,
        "profile_url": profile_url,
    }


# Mapped from current Google Business account (confirmed in manager UI).
# Key = row number from Yandex sheet.
GOOGLE_BY_ROW = {
    5: [
        listing(
            name="Сырная Лавка",
            address="879P+G5P, Tashkent, Uzbekistan",
            status="Processing",
            fid="13466168350756228991",
            store_code="04732619173853016988",
            profile_url="https://business.google.com/n/5596989586233886836/profile?fid=13466168350756228991",
        )
    ],
    8: [
        listing(
            name="Syrnaya Lavka Kuylyuk",
            address="68PJ+H5R, Tashkent",
            status="Processing",
            fid="2943381179443995567",
            store_code="04773736263072149968",
            profile_url="https://business.google.com/n/5659155215486045136/profile?fid=2943381179443995567",
        )
    ],
    15: [
        listing(
            name="Сырная Лавка",
            address="Near yakkasaroy street, Tashkent",
            status="Verified",
            fid="12242063773800875924",
            store_code="",
            profile_url="https://business.google.com/n/2036140660410611774/profile?fid=12242063773800875924",
        )
    ],
    25: [
        listing(
            name="Syrnaya Lavka",
            address="978R+2Q7, Yangi Shahar ko'chasi, 100194, Tashkent",
            status="Verified",
            fid="10854957917197044563",
            store_code="04445343178342222593",
            profile_url="https://business.google.com/n/15283971431114310555/profile?fid=10854957917197044563",
        ),
        listing(
            name="Syrnaya Lavka",
            address="Abdurauf Fitrat ko'chasi 4, Tashkent",
            status="Processing",
            fid="7022877837922159990",
            store_code="01083990800331106501",
            profile_url="https://business.google.com/n/18418561246523130052/profile?fid=7022877837922159990",
        ),
    ],
    20: [
        listing(
            name="Сырная Лавка",
            address="Tansiqboev ko'chasi 25, Tashkent",
            status="Processing",
            fid="16639199160416746131",
            store_code="11569504432881214481",
            profile_url="https://business.google.com/n/11234282776194258137/profile?fid=16639199160416746131",
        )
    ],
    28: [
        listing(
            name="Syrnaya Lavka Ttz",
            address="993P+JFQ Ттз базар, Тоshkent",
            status="Processing",
            fid="16287832862762180124",
            store_code="13269299752577284755",
            profile_url="https://business.google.com/n/15475703256441063774/profile?fid=16287832862762180124",
        )
    ],
}


def load_last_attempts() -> dict[int, dict]:
    if not ATTEMPTS_PATH.exists():
        return {}
    try:
        data = json.loads(ATTEMPTS_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}
    last_by_row: dict[int, dict] = {}
    for item in data:
        row = item.get("row")
        if isinstance(row, int):
            last_by_row[row] = item
    return last_by_row


def main():
    wb = load_workbook(EXCEL_PATH)
    source = wb[wb.sheetnames[0]]

    if "google map" in wb.sheetnames:
        del wb["google map"]
    ws = wb.create_sheet("google map")

    headers = [
        "#",
        "Название в Google",
        "Внутренний ID\n(заполнить)",
        "Адрес (Google)",
        "Ссылка Google Business",
        "Время работы",
        "Дни работы",
        "Телефон",
        "Google Profile ID",
        "Статус / Разница",
        "Рейтинг Google (оценка/голоса)",
        "Что не хватает",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(1, i).value = h

    row_kind = {}
    last_attempts = load_last_attempts()
    out_row = 2

    for r in range(2, source.max_row + 1):
        row_no = source.cell(r, 1).value
        if not isinstance(row_no, (int, float)):
            continue
        row_no = int(row_no)

        brand_excel = str(source.cell(r, 2).value or "").strip()
        internal_id = source.cell(r, 3).value or ""
        hours_expected = source.cell(r, 6).value or ""
        days_expected = source.cell(r, 7).value or ""

        listings = GOOGLE_BY_ROW.get(row_no, [])
        primary = listings[0] if listings else None

        if primary:
            name_google = primary["name"]
            address_google = primary["address"]
            link_google = primary["profile_url"]
            ids = []
            for x in listings:
                id_bits = [f"fid:{x['fid']}"]
                if x["store_code"]:
                    id_bits.append(f"code:{x['store_code']}")
                ids.append(" ".join(id_bits))
            profile_ids = " | ".join(ids)

            statuses = [x["status"] for x in listings]
            if "Verified" in statuses and "Processing" in statuses:
                status_text = "есть в Google (Verified) + дубль Processing"
                action_text = "проверить и объединить дубли после завершения Processing"
                kind = "have_check"
            elif "Processing" in statuses:
                status_text = "claim отправлен, Google Processing"
                action_text = "ждём завершения проверки Google"
                kind = "pending"
            else:
                status_text = "есть в Google (Verified)"
                action_text = "проверить телефон/часы/описание и заполнить рейтинг"
                kind = "have"

            phone = "+998 78 555 15 15"
            rating = ""
        else:
            name_google = ""
            address_google = ""
            link_google = ""
            profile_ids = ""
            rating = ""
            phone = ""

            if brand_excel.upper() == "RUBA":
                status_text = "ждём информацию (RUBA)"
                action_text = "создать/найти RUBA карточку в Google отдельно от Сырная Лавка"
                kind = "ruba"
            else:
                status_text = "ждём информацию: не найдено в нашем Google Business"
                action_text = "нужно найти карточку в Google Maps и нажать Claim this business"
                kind = "waiting"
                attempt = last_attempts.get(row_no, {})
                result = str(attempt.get("result") or "")
                if result == "rpc_error_persists":
                    action_text = "RpcError again: retry in clean session, then open Google Support ticket"
                elif result == "not_found_search":
                    action_text = "Not found by search: try plus-code/pin, else create a new listing"
                elif result == "ambiguous_place":
                    action_text = "Ambiguous match: need exact pin/share-link from owner"
                elif result == "blocked_address_validation":
                    action_text = "Address validation blocked in claim wizard: support-assisted claim needed"
                elif result == "no_claim_cta":
                    action_text = "No Claim/Manage CTA: try direct listing URL or alternate account"
                elif result in {"already_managed_match", "duplicate_of_managed"}:
                    status_text = "already covered by managed Google profile"
                    action_text = "Verify row-to-profile mapping and mark as covered"
                    kind = "have_check"

        ws.cell(out_row, 1).value = row_no
        ws.cell(out_row, 2).value = name_google
        ws.cell(out_row, 3).value = internal_id
        ws.cell(out_row, 4).value = address_google
        ws.cell(out_row, 5).value = link_google
        ws.cell(out_row, 6).value = hours_expected
        ws.cell(out_row, 7).value = days_expected
        ws.cell(out_row, 8).value = phone
        ws.cell(out_row, 9).value = profile_ids
        ws.cell(out_row, 10).value = status_text
        ws.cell(out_row, 11).value = rating
        ws.cell(out_row, 12).value = action_text
        row_kind[out_row] = kind
        out_row += 1

    max_row = out_row - 1
    max_col = 12

    for i, col in enumerate("ABCDEFGHIJKL", start=1):
        ws.column_dimensions[col].width = source.column_dimensions[col].width

    for r in range(1, max_row + 1):
        ws.row_dimensions[r].height = source.row_dimensions[r].height
        for c in range(1, max_col + 1):
            src_cell = source.cell(r, c)
            dst_cell = ws.cell(r, c)
            dst_cell._style = copy(src_cell._style)

    for c in range(10, 13):
        ws.cell(1, c)._style = copy(source.cell(1, 1)._style)

    status_style = copy(source.cell(3, 11)._style)
    status_visuals = {
        "have": {"fill": solid_fill("FFC6EFCE"), "font": color_font("FF006100")},
        "have_check": {"fill": solid_fill("FFFFF2CC"), "font": color_font("FF7F6000")},
        "pending": {"fill": solid_fill("FFFCE4D6"), "font": color_font("FF9C5700")},
        "waiting": {"fill": solid_fill("FFF8CBAD"), "font": color_font("FF9C0006")},
        "ruba": {"fill": solid_fill("FFD9E1F2"), "font": color_font("FF1F4E78")},
    }

    for r in range(2, max_row + 1):
        ws.cell(r, 10)._style = copy(status_style)
        visual = status_visuals.get(row_kind.get(r))
        if visual:
            ws.cell(r, 10).fill = copy(visual["fill"])
            ws.cell(r, 10).font = copy(visual["font"])

    for r in range(2, max_row + 1):
        link = ws.cell(r, 5).value
        if link:
            ws.cell(r, 5).hyperlink = str(link)

    ws.auto_filter.ref = f"A1:L{max_row}"
    ws.freeze_panes = "A2"
    wb.save(EXCEL_PATH)

    kind_counts = {}
    for v in row_kind.values():
        kind_counts[v] = kind_counts.get(v, 0) + 1
    stats = {
        "rows_total": max_row - 1,
        "have_verified": kind_counts.get("have", 0),
        "have_pending": kind_counts.get("pending", 0),
        "have_check": kind_counts.get("have_check", 0),
        "waiting_info": kind_counts.get("waiting", 0),
        "waiting_info_ruba": kind_counts.get("ruba", 0),
        "kind_counts": kind_counts,
    }
    print(json.dumps(stats, ensure_ascii=False))


if __name__ == "__main__":
    main()
