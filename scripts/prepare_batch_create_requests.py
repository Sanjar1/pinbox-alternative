#!/usr/bin/env python3
import csv
import json
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
DOCS_DIR = ROOT / "docs"
EXCEL_PATH = DATA_DIR / "stores_audit.xlsx"

PENDING_MARKERS = [
    "pending",
    "processing",
    "moderation",
    "модерац",
    "на провер",
    "в обработ",
    "ожидан",
    "заявка",
    "отправлен",
]


def as_text(v):
    return "" if v is None else str(v).strip()


def is_pending_status(status_text: str) -> bool:
    t = status_text.lower()
    return any(marker in t for marker in PENDING_MARKERS)


def sheet_rows(ws):
    # Known column layout in current workbook
    # 1:# 2:name 3:internal_id 4:address 5:url 6:hours 7:days 8:phone 9:platform_id 10:status
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r:
            continue
        row_num = r[0]
        if row_num in (None, ""):
            continue
        if not isinstance(row_num, (int, float)):
            continue
        yield {
            "row": int(row_num),
            "name": as_text(r[1]) if len(r) > 1 else "",
            "internal_id": as_text(r[2]) if len(r) > 2 else "",
            "address": as_text(r[3]) if len(r) > 3 else "",
            "map_url": as_text(r[4]) if len(r) > 4 else "",
            "hours": as_text(r[5]) if len(r) > 5 else "",
            "days": as_text(r[6]) if len(r) > 6 else "",
            "phone": as_text(r[7]) if len(r) > 7 else "",
            "platform_id": as_text(r[8]) if len(r) > 8 else "",
            "status": as_text(r[9]) if len(r) > 9 else "",
        }


def build_requests():
    wb = load_workbook(EXCEL_PATH, data_only=True)
    platform_map = {
        "Yandex": "yandex",
        "2gis": "2gis",
        "google map": "google",
    }

    results = {"generated_at": date.today().isoformat(), "platforms": {}}

    for sheet_name, platform in platform_map.items():
        ws = wb[sheet_name]
        missing = []
        pending = []
        have_id = 0
        total = 0

        for row in sheet_rows(ws):
            total += 1
            has_id = row["platform_id"].isdigit()
            pending_flag = is_pending_status(row["status"])

            entry = {
                "platform": platform,
                "sheet": sheet_name,
                "row": row["row"],
                "store_name": row["name"],
                "internal_id": row["internal_id"],
                "address": row["address"],
                "map_url": row["map_url"],
                "phone": row["phone"],
                "hours": row["hours"],
                "days": row["days"],
                "platform_id": row["platform_id"] or None,
                "status": row["status"] or None,
                "action": None,
            }

            if has_id:
                have_id += 1
                continue

            if pending_flag:
                entry["action"] = "wait_pending"
                pending.append(entry)
            else:
                entry["action"] = "create_request"
                missing.append(entry)

        results["platforms"][platform] = {
            "sheet": sheet_name,
            "total_rows": total,
            "with_id": have_id,
            "pending_without_id": len(pending),
            "create_request_needed": len(missing),
            "pending_items": pending,
            "create_items": missing,
        }

    return results


def write_outputs(payload):
    out_json = DATA_DIR / f"platform_create_requests_{payload['generated_at']}.json"
    out_csv = DATA_DIR / f"platform_create_requests_{payload['generated_at']}.csv"

    out_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    rows = []
    for platform_data in payload["platforms"].values():
        rows.extend(platform_data["create_items"])

    with out_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "platform",
                "sheet",
                "row",
                "store_name",
                "internal_id",
                "address",
                "map_url",
                "phone",
                "hours",
                "days",
                "platform_id",
                "status",
                "action",
            ],
        )
        w.writeheader()
        for r in rows:
            w.writerow(r)

    doc_path = DOCS_DIR / f"BATCH_CREATE_REQUESTS_{payload['generated_at']}.md"

    y = payload["platforms"]["yandex"]
    d2 = payload["platforms"]["2gis"]
    g = payload["platforms"]["google"]

    doc = f"""# Batch Create Requests ({payload['generated_at']})

## Summary

- Yandex: create={y['create_request_needed']}, pending={y['pending_without_id']}, with_id={y['with_id']}/{y['total_rows']}
- 2GIS: create={d2['create_request_needed']}, pending={d2['pending_without_id']}, with_id={d2['with_id']}/{d2['total_rows']}
- Google: create={g['create_request_needed']}, pending={g['pending_without_id']}, with_id={g['with_id']}/{g['total_rows']}

## Output files

- `{out_json.relative_to(ROOT)}`
- `{out_csv.relative_to(ROOT)}`

## API-first policy (batch changes)

- Use official/private API endpoints first for all bulk operations.
- Use browser UI only for flows without stable API coverage (for example: ownership claims and some listing creation flows).
- Keep a queue file (`data/platform_create_requests_*.json`) and execute in batches.
- Log each run and keep result snapshots in `data/`.

## Platform notes

- Yandex: no fully stable public listing-create API confirmed; creation is usually via `/sprav/add` UI flow.
- 2GIS: private cabinet APIs are available for updates; creation endpoint for new org/branch is not confirmed in current capture.
- Google: Business Profile APIs support location create/update/review reply when account is authorized.

## Next execution step

1. Process all `action=create_request` rows from CSV.
2. For Yandex/2GIS without create endpoint, submit via cabinet UI in batch sessions and update IDs in Excel.
3. After IDs appear, upload `data/roba_logo_only.jpg` for RUBA rows.
"""

    doc_path.write_text(doc, encoding="utf-8")
    return out_json, out_csv, doc_path


def main():
    payload = build_requests()
    out_json, out_csv, doc_path = write_outputs(payload)
    print(json.dumps({
        "json": str(out_json),
        "csv": str(out_csv),
        "doc": str(doc_path),
        "summary": {
            k: {
                "create": v["create_request_needed"],
                "pending": v["pending_without_id"],
                "with_id": v["with_id"],
                "total": v["total_rows"],
            }
            for k, v in payload["platforms"].items()
        }
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
