#!/usr/bin/env python3
"""Sync Telegram progress JSON into the Yandex sheet in stores_audit.xlsx."""

from __future__ import annotations

import argparse
import json
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
DEFAULT_XLSX = DATA_DIR / "stores_audit.xlsx"


def _latest_telegram_status_file() -> Path:
    candidates = sorted(DATA_DIR.glob("stores_telegram_status_*.json"))
    if not candidates:
        raise FileNotFoundError("No stores_telegram_status_*.json files found in data/")
    return candidates[-1]


def _load_status(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def _to_int_set(values: list[Any]) -> set[int]:
    out: set[int] = set()
    for value in values:
        try:
            out.add(int(value))
        except (TypeError, ValueError):
            continue
    return out


def sync_tracking(source_xlsx_path: Path, output_xlsx_path: Path, status_path: Path) -> dict[str, Any]:
    status = _load_status(status_path)

    telegram_url = status.get("telegram_url", "https://t.me/SirnayaLavka_Uzb")
    completed_ids = _to_int_set([item.get("id") for item in status.get("completed_stores", [])])
    pending_ids = _to_int_set(status.get("pending_stores", []))
    accessible_ids = completed_ids | pending_ids

    wb = openpyxl.load_workbook(source_xlsx_path)
    ws = wb["Yandex"]

    headers = {
        15: "Telegram URL",
        16: "Telegram Status",
        17: "Amenities %",
        18: "Rating Stars",
        19: "Review Count",
        20: "Overall %",
        21: "Last Updated",
    }
    for col, title in headers.items():
        ws.cell(row=1, column=col).value = title

    done = 0
    pending = 0
    needs_claim = 0

    for row in range(2, ws.max_row + 1):
        org_id_value = ws.cell(row=row, column=9).value
        if org_id_value in (None, ""):
            continue

        try:
            org_id = int(org_id_value)
        except (TypeError, ValueError):
            continue

        ws.cell(row=row, column=15).value = telegram_url
        ws.cell(row=row, column=17).value = "Pending"
        ws.cell(row=row, column=18).value = "Pending"
        ws.cell(row=row, column=19).value = "Pending"

        if org_id in completed_ids:
            status_text = "Done"
            overall = 33
            done += 1
        elif org_id in pending_ids:
            status_text = "Pending"
            overall = 0
            pending += 1
        elif org_id in accessible_ids:
            status_text = "Pending"
            overall = 0
            pending += 1
        else:
            status_text = "Needs claim"
            overall = 0
            needs_claim += 1

        ws.cell(row=row, column=16).value = status_text
        ws.cell(row=row, column=20).value = overall
        ws.cell(row=row, column=21).value = status.get("date") or date.today().isoformat()

    wb.save(output_xlsx_path)

    summary = {
        "date": status.get("date") or date.today().isoformat(),
        "source": str(status_path.relative_to(ROOT)),
        "source_xlsx": str(source_xlsx_path.relative_to(ROOT)),
        "output_xlsx": str(output_xlsx_path.relative_to(ROOT)),
        "telegram_url": telegram_url,
        "accessible_total": len(accessible_ids),
        "telegram_done": done,
        "telegram_pending": pending,
        "needs_claim": needs_claim,
    }
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--output", type=Path, default=DATA_DIR / "stores_audit_tracking_2026-02-28.xlsx")
    parser.add_argument("--status", type=Path, default=None)
    parser.add_argument("--summary", type=Path, default=DATA_DIR / "yandex_tracking_sync_summary_2026-02-28.json")
    args = parser.parse_args()

    status_path = args.status or _latest_telegram_status_file()
    source_xlsx_path = args.xlsx
    output_xlsx_path = args.output

    if not status_path.exists():
        raise FileNotFoundError(f"Status file not found: {status_path}")
    if not source_xlsx_path.exists():
        raise FileNotFoundError(f"Workbook not found: {source_xlsx_path}")

    summary = sync_tracking(
        source_xlsx_path=source_xlsx_path,
        output_xlsx_path=output_xlsx_path,
        status_path=status_path,
    )
    args.summary.parent.mkdir(parents=True, exist_ok=True)
    with args.summary.open("w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
