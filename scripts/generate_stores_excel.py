"""
Generate stores audit Excel file — Сырная Лавка (33 stores).
Data collected from Yandex Business via browser session.
Run: python scripts/generate_stores_excel.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

P78 = "+998 78 555 15 15"
P97 = "+998 97 711 15 15"

STORES = [
    (2605231525,  "08:00-22:00", "Ежедневно",  f"{P78}, {P97}",           "Ташкент, Чиланзарский район, махаллинский сход граждан Катта Чиланзар-3"),
    (46711213257, "08:00-20:00", "Ежедневно",  f"{P78}, {P97}",           "Ташкент, Яшнабадский район, махаллинский сход граждан Парвоз"),
    (51521899757, "08:00-20:00", "Ежедневно",  P78,                        "Ташкент, Мирзо-Улугбекский район, массив Карасу, 6-й квартал, 28Г"),
    (68372174039, "08:00-19:00", "Ежедневно",  f"{P78}, {P97}",           "Ташкент, улица Сакичмон, 1"),
    (73077844158, "08:00-20:00", "Ежедневно",  P78,                        "Ташкент, проспект Амира Темура, 42/14"),
    (78130811373, "08:00-20:00", "Ежедневно",  P78,                        "Ташкент, улица Паркент, 74"),
    (80285992156, "08:00-20:00", "Ежедневно",  f"{P97}, {P78}",           "Ташкент, Мирзо-Улугбекский район, улица Сайрам, 37"),
    (81444134916, "08:00-20:00", "⚠ weekend",  P78,                        "Ташкент, Большая кольцевая дорога, 17/10"),
    (88969661261, "08:00-20:00", "Ежедневно",  P78,                        "Ташкент, улица Катартал, 60"),
    (93021421517, "08:00-20:00", "Ежедневно",  f"{P97}, {P78}",           "Ташкент, улица Мирабад, 43"),
    (93653304255, "07:00-19:00", "Ежедневно",  P78,                        "Ташкент, Бектемирский район, мсг Нурлийул, улица Мехнатабад, 82"),
    (96275437524, "08:00-20:00", "Ежедневно",  f"{P78}, {P97}",           "Ташкент, Яшнабадский район, махаллинский сход граждан Парвоз"),
    (98808908571, "08:00-20:00", "Ежедневно",  P78,                        "Ташкент, Сергелийский район, мсг Мехригиё"),
    (113993963061,"08:00-20:00", "Ежедневно",  f"{P78}, +998 90 372 04 40","Ташкент, Яшнабадский район, мсг Уйсозлар"),
    (119087534313,"08:00-19:00", "Ежедневно",  P78,                        "Ташкент, улица Катартал, 60А/1"),
    (119523779091,"08:00-20:00", "Ежедневно",  f"{P97}, {P78}",           "Ташкент, Яккасарайский район, мсг Шахжахан"),
    (133701096811,"08:00-17:00", "⚠ saturday", P78,                        "Ташкент, улица Уйсозлар, 49"),
    (134404129580,"08:00-20:00", "⚠ weekend",  f"{P78}, {P97}",           "Ташкент, Фархадский базар"),
    (140717986697,"08:00-20:00", "Ежедневно",  P78,                        "Ташкент, Юнусабадский район, массив Юнусабад, 3-й квартал, 3/6"),
    (140997774388,"08:00-19:00", "Ежедневно",  f"{P97}, {P78}",           "Ташкент, Мирзо-Улугбекский район, мсг Хабиб Абдуллаев"),
    (143672341206,"10:00-19:00", "Ежедневно",  P78,                        "Ташкентская область, Бостанлыкский район, Газалкент"),
    (146603754824,"09:00-18:00", "⚠ weekdays", P78,                        "Ташкент, Мирзо-Улугбекский район, массив Ялангач, 30В"),
    (160095672246,"08:00-20:00", "Ежедневно",  f"{P78}, {P97}",           "Ташкент, Мирзо-Улугбекский район, мсг Буюк Ипак Йули"),
    (168675219928,"08:00-20:00", "Ежедневно",  P78,                        "Ташкент, Алмазарский район, массив Бешкурган, 1-й квартал"),
    (189015626941,"08:00-20:00", "Ежедневно",  P78,                        "Ташкент, Юнусабадский район, мсг Акбарабад"),
    (191697629628,"08:00-19:00", "Ежедневно",  P78,                        "Ташкент, Чиланзарский район, мсг Навбахор"),
    (193938967033,"08:00-20:00", "Ежедневно",  P78,                        "Ташкент, улица Тимура Малика, 3"),
    (205196568796,"—",           "⚠ не задано", P78,                       "Ташкент, Учтепинский район, мсг Бирлик"),
    (219043654252,"08:00-20:00", "Ежедневно",  P78,                        "Ташкент, Юнусабадский район, массив Юнусабад, 14-й квартал, 61А"),
    (225503578112,"08:00-20:00", "Ежедневно",  P78,                        "Ташкент, улица Абдурауфа Фитрата, 4"),
    (225833833825,"08:00-20:00", "Ежедневно",  P78,                        "Ташкентская область, Чирчик"),
    (235345012305,"08:00-22:00", "Ежедневно",  P78,                        "Ташкент, улица Карасу, 1"),
    (242380255215,"08:00-20:00", "Ежедневно",  P78,                        "Ташкент, Мирзо-Улугбекский район, мсг Ахиллик"),
]

def create_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Магазины"

    # Styles
    hdr_fill   = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
    hdr_font   = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    alt_fill   = PatternFill(start_color="FFF8F0", end_color="FFF8F0", fill_type="solid")
    warn_fill  = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    link_font  = Font(color="0563C1", underline="single", name="Calibri")
    body_font  = Font(name="Calibri", size=10)
    center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin       = Side(style="thin", color="E0E0E0")
    bdr        = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Column config: (letter, header text, width)
    COLS = [
        ("A", "#",                      5),
        ("B", "Название в Яндекс",     22),
        ("C", "Внутренний ID\n(заполнить)", 18),
        ("D", "Адрес (Яндекс)",        42),
        ("E", "Ссылка Яндекс.Карты",   38),
        ("F", "Время работы",          15),
        ("G", "Дни работы",            18),
        ("H", "Телефон",               32),
        ("H", "Телефон",               32),  # placeholder
    ]
    # Fix: define properly
    cols = [
        ("A", "#",                      5),
        ("B", "Название в Яндекс",     22),
        ("C", "Внутренний ID\n(заполнить)", 18),
        ("D", "Адрес (Яндекс)",        42),
        ("E", "Ссылка Яндекс.Карты",   38),
        ("F", "Время работы",          15),
        ("G", "Дни работы",            18),
        ("H", "Телефон",               32),
        ("I", "Yandex Org ID",         18),
    ]

    for col_l, title, width in cols:
        cell = ws[f"{col_l}1"]
        cell.value = title
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = bdr
        ws.column_dimensions[col_l].width = width

    ws.row_dimensions[1].height = 36

    for i, (org_id, hours, days, phones, address) in enumerate(STORES, start=1):
        row = i + 1
        is_alt   = (i % 2 == 0)
        has_warn = "⚠" in days
        bg = warn_fill if has_warn else (alt_fill if is_alt else None)
        map_url  = f"https://yandex.ru/maps/org/{org_id}"

        def st(cell, align=left_wrap):
            if bg: cell.fill = bg
            cell.font = body_font
            cell.alignment = align
            cell.border = bdr

        c = ws[f"A{row}"]; c.value = i;            st(c, center)
        c = ws[f"B{row}"]; c.value = "Сырная Лавка"; st(c)
        c = ws[f"C{row}"]; c.value = "";             st(c, center)  # user fills
        c = ws[f"D{row}"]; c.value = address;        st(c)
        c = ws[f"E{row}"]; c.value = map_url; c.hyperlink = map_url; c.font = link_font; c.alignment = left_wrap; c.border = bdr
        if bg: c.fill = bg
        c = ws[f"F{row}"]; c.value = hours;          st(c, center)
        c = ws[f"G{row}"]; c.value = days;           st(c, center)
        c = ws[f"H{row}"]; c.value = phones;         st(c)
        c = ws[f"I{row}"]; c.value = org_id;         st(c, center)

        ws.row_dimensions[row].height = 22

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(STORES)+1}"

    # Legend row
    legend_row = len(STORES) + 3
    ws[f"A{legend_row}"].value = "⚠ = проверить расписание в Яндексе (weekend/weekdays/saturday — нестандартное расписание)"
    ws[f"A{legend_row}"].font = Font(italic=True, color="856404", name="Calibri")

    out_path = "data/stores_audit.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path} - {len(STORES)} stores")
    print(f"  Stores needing schedule check: {sum(1 for _,_,d,_,_ in STORES if d.startswith('?'))}")

if __name__ == "__main__":
    create_excel()
