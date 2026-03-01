import json
import math
import re
import time
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


TOKEN = "9ef36758fb840f45fd63ff5ca9ddf5c53feb4564"
ORG_ID = "70000001036827089"
CATALOG_KEY = "rupycp2722"
EXCEL_PATH = Path("data/stores_audit.xlsx")
REPORT_PATH = Path("data/2gis_coordinate_sync_report.json")
UPDATES_PATH = Path("data/2gis_hours_updates_result.json")


def as_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_hours(value):
    text = as_text(value).replace(" ", "")
    text = text.replace("–", "-").replace("—", "-")
    m = re.match(r"^(\d{2}:\d{2})-(\d{2}:\d{2})$", text)
    if not m:
        return None
    return f"{m.group(1)}-{m.group(2)}"


def extract_org_id(url, fallback):
    m = re.search(r"/org/(\d+)", as_text(url))
    if m:
        return m.group(1)
    fallback_text = as_text(fallback)
    if fallback_text.isdigit():
        return fallback_text
    return None


def parse_float(value):
    text = as_text(value).replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_point_from_url(url):
    text = as_text(url)
    if not text or text in ("-", "—"):
        return None
    parsed = urllib.parse.urlparse(text)
    params = urllib.parse.parse_qs(parsed.query)

    def parse_lon_lat(value):
        parts = value.split(",")
        if len(parts) != 2:
            return None
        lon = parse_float(parts[0])
        lat = parse_float(parts[1])
        if lon is None or lat is None:
            return None
        return {"lat": lat, "lon": lon}

    for key in ("ll", "whatshere[point]", "pt"):
        vals = params.get(key)
        if vals:
            point = parse_lon_lat(vals[0])
            if point:
                return point

    m = re.search(r"ll=([\-0-9.]+),([\-0-9.]+)", text)
    if m:
        return {"lat": float(m.group(2)), "lon": float(m.group(1))}
    return None


def parse_yandex_coordinates(html):
    patterns = [
        r"coordinates:\[([\-0-9.]+),([\-0-9.]+)\]",
        r'"coordinates":\[([\-0-9.]+),([\-0-9.]+)\]',
        r"\[([\-0-9.]+),([\-0-9.]+)\],\s*\"type\":\"Point\"",
    ]
    for pattern in patterns:
        m = re.search(pattern, html)
        if m:
            lon = float(m.group(1))
            lat = float(m.group(2))
            return {"lat": lat, "lon": lon}
    return None


def get_text(url, headers=None):
    req = urllib.request.Request(url, headers=headers or {})
    with urllib.request.urlopen(req, timeout=40) as resp:
        return resp.read().decode("utf-8", errors="replace")


def get_json(url, headers=None, method="GET", body=None):
    req = urllib.request.Request(
        url,
        data=body,
        headers=headers or {},
        method=method,
    )
    with urllib.request.urlopen(req, timeout=40) as resp:
        return json.loads(resp.read().decode("utf-8"))


def read_excel_rows():
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    headers = [as_text(cell.value).lower() for cell in ws[1]]
    lat_idx = None
    lon_idx = None
    for idx, h in enumerate(headers):
        if lat_idx is None and ("lat" in h or "широт" in h):
            lat_idx = idx
        if lon_idx is None and ("lon" in h or "lng" in h or "долгот" in h):
            lon_idx = idx

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not any(v not in (None, "", " ", "-", "—") for v in r):
            continue
        row_num = r[0]
        row_id = int(row_num) if isinstance(row_num, (int, float)) else None
        if row_id is None:
            continue
        yandex_url = as_text(r[4]) if len(r) > 4 else ""
        if yandex_url in ("-", "—"):
            yandex_url = ""
        manual_lat = parse_float(r[lat_idx]) if lat_idx is not None and lat_idx < len(r) else None
        manual_lon = parse_float(r[lon_idx]) if lon_idx is not None and lon_idx < len(r) else None
        manual_point = None
        if manual_lat is not None and manual_lon is not None:
            manual_point = {"lat": manual_lat, "lon": manual_lon}
        url_point = parse_point_from_url(yandex_url)
        point = manual_point or url_point
        point_source = "manual_columns" if manual_point else ("url_params" if url_point else None)
        rows.append(
            {
                "row": row_id,
                "brand": as_text(r[1]),
                "internal_id": as_text(r[2]),
                "address": as_text(r[3]),
                "yandex_url": yandex_url,
                "hours": normalize_hours(r[5]),
                "days": as_text(r[6]),
                "phone": as_text(r[7]),
                "yandex_org_id": extract_org_id(yandex_url, r[8]),
                "yandex_point": point,
                "yandex_point_source": point_source,
            }
        )
    return rows


def get_yandex_points(rows):
    cache = {}
    failed = []
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
        )
    }
    for row in rows:
        if row.get("yandex_point"):
            continue
        org_id = row["yandex_org_id"]
        if not org_id:
            failed.append({"row": row["row"], "reason": "no_yandex_org_id_or_point"})
            continue
        if org_id in cache:
            row["yandex_point"] = cache[org_id]
            row["yandex_point_source"] = "org_cache"
            continue
        try:
            html = get_text(f"https://yandex.ru/maps/org/{org_id}", headers=headers)
            if "showcaptcha" in html or "smart-captcha" in html or "__SSR_DATA__={url:\"/ru/checkbox" in html:
                cache[org_id] = None
                row["yandex_point"] = None
                row["yandex_point_source"] = None
                failed.append({"row": row["row"], "org_id": org_id, "reason": "captcha_blocked"})
                time.sleep(0.3)
                continue
            point = parse_yandex_coordinates(html)
            cache[org_id] = point
            row["yandex_point"] = point
            row["yandex_point_source"] = "org_page" if point else None
            if not point:
                failed.append({"row": row["row"], "org_id": org_id, "reason": "coordinates_not_found"})
        except Exception as exc:
            cache[org_id] = None
            row["yandex_point"] = None
            failed.append({"row": row["row"], "org_id": org_id, "reason": f"fetch_error: {exc}"})
        time.sleep(0.3)
    return failed


def parse_branch_daily_hours(schedule):
    if not isinstance(schedule, dict):
        return None
    days = schedule.get("days")
    if not isinstance(days, dict):
        return None
    vals = set()
    for day in days.values():
        if isinstance(day, dict):
            fr = day.get("from")
            to = day.get("to")
            if fr and to:
                vals.add((fr, to))
    if len(vals) != 1:
        return None
    fr, to = next(iter(vals))
    return f"{fr}-{to}"


def load_2gis_branches():
    headers = {"Authorization": f"Bearer {TOKEN}"}
    page = 1
    page_size = 50
    result = []
    while True:
        url = (
            f"https://api.account.2gis.com/api/1.0/branches"
            f"?orgId={ORG_ID}&fields=schedule&pageSize={page_size}&page={page}"
        )
        payload = get_json(url, headers=headers)
        items = payload.get("result", {}).get("items", [])
        if not items:
            break
        result.extend(items)
        if len(items) < page_size:
            break
        page += 1
    return result


def get_branch_point(branch_id):
    url = (
        "https://catalog.api.2gis.com/3.0/items/byid"
        f"?id={branch_id}&fields=items.point,items.id&key={CATALOG_KEY}"
    )
    payload = get_json(url)
    items = payload.get("result", {}).get("items", [])
    if not items:
        return None
    point = items[0].get("point")
    if not isinstance(point, dict):
        return None
    lat = point.get("lat")
    lon = point.get("lon")
    if lat is None or lon is None:
        return None
    return {"lat": float(lat), "lon": float(lon)}


def enrich_branch_points(branches):
    cache = {}
    failures = []
    for b in branches:
        bid = str(b.get("id", ""))
        point = None
        if bid.isdigit():
            if bid in cache:
                point = cache[bid]
            else:
                try:
                    point = get_branch_point(bid)
                    cache[bid] = point
                except Exception as exc:
                    cache[bid] = None
                    failures.append({"branch_id": bid, "reason": f"point_fetch_error: {exc}"})
                time.sleep(0.15)
        b["point"] = point
        b["daily_hours"] = parse_branch_daily_hours(b.get("schedule"))
    return failures


def haversine_m(lat1, lon1, lat2, lon2):
    r = 6371000
    p1 = math.radians(lat1)
    p2 = math.radians(lat2)
    d_phi = math.radians(lat2 - lat1)
    d_lam = math.radians(lon2 - lon1)
    a = math.sin(d_phi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(d_lam / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


def assign_by_distance(rows, branches, threshold_m=450):
    indexed_rows = [r for r in rows if r.get("yandex_point")]
    indexed_branches = [b for b in branches if b.get("point")]
    candidates = []
    for r in indexed_rows:
        rp = r["yandex_point"]
        for b in indexed_branches:
            bp = b["point"]
            dist = haversine_m(rp["lat"], rp["lon"], bp["lat"], bp["lon"])
            candidates.append((dist, r, b))
    candidates.sort(key=lambda x: x[0])

    assigned_rows = set()
    assigned_branches = set()
    matches = []
    for dist, r, b in candidates:
        row_key = r["row"] if r["row"] is not None else id(r)
        branch_key = str(b.get("id"))
        if row_key in assigned_rows or branch_key in assigned_branches:
            continue
        if dist > threshold_m:
            continue
        assigned_rows.add(row_key)
        assigned_branches.add(branch_key)
        hour_mismatch = bool(r.get("hours") and b.get("daily_hours") and r["hours"] != b["daily_hours"])
        matches.append(
            {
                "row": r["row"],
                "brand_excel": r["brand"],
                "internal_id": r["internal_id"],
                "hours_excel": r["hours"],
                "address_excel": r["address"],
                "yandex_org_id": r["yandex_org_id"],
                "yandex_point": r["yandex_point"],
                "branch_id": str(b.get("id")),
                "branch_name": b.get("name"),
                "branch_address": b.get("address"),
                "branch_point": b.get("point"),
                "branch_daily_hours": b.get("daily_hours"),
                "distance_m": round(dist, 2),
                "hour_mismatch": hour_mismatch,
            }
        )

    matched_row_ids = {m["row"] for m in matches}
    unmatched_rows = []
    for r in rows:
        if not r.get("yandex_point"):
            unmatched_rows.append(
                {
                    "row": r["row"],
                    "brand_excel": r["brand"],
                    "internal_id": r["internal_id"],
                    "reason": "no_yandex_point",
                    "yandex_org_id": r["yandex_org_id"],
                }
            )
            continue
        if r["row"] not in matched_row_ids:
            rp = r["yandex_point"]
            nearest = None
            for b in indexed_branches:
                bp = b["point"]
                dist = haversine_m(rp["lat"], rp["lon"], bp["lat"], bp["lon"])
                if nearest is None or dist < nearest[0]:
                    nearest = (dist, b)
            unmatched_rows.append(
                {
                    "row": r["row"],
                    "brand_excel": r["brand"],
                    "internal_id": r["internal_id"],
                    "reason": "distance_too_far_or_branch_already_used",
                    "nearest_branch_id": str(nearest[1].get("id")) if nearest else None,
                    "nearest_branch_address": nearest[1].get("address") if nearest else None,
                    "nearest_distance_m": round(nearest[0], 2) if nearest else None,
                }
            )

    matched_branch_ids = {m["branch_id"] for m in matches}
    unmatched_branches = [
        {
            "branch_id": str(b.get("id")),
            "branch_name": b.get("name"),
            "branch_address": b.get("address"),
            "point": b.get("point"),
            "daily_hours": b.get("daily_hours"),
        }
        for b in branches
        if str(b.get("id")) not in matched_branch_ids
    ]
    return matches, unmatched_rows, unmatched_branches


def build_schedule(hours):
    fr, to = hours.split("-")
    days = {}
    for d in ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]:
        days[d] = {"from": fr, "to": to}
    return {"days": days, "comment": None}


def update_branch_hours(updates):
    headers = {"Authorization": f"Bearer {TOKEN}", "Content-Type": "application/json"}
    results = []
    for upd in updates:
        branch_id = upd["branch_id"]
        body = {"schedule": build_schedule(upd["hours_excel"])}
        try:
            payload = json.dumps(body, ensure_ascii=False).encode("utf-8")
            response = get_json(
                f"https://api.account.2gis.com/api/1.0/branches/{branch_id}",
                headers=headers,
                method="PUT",
                body=payload,
            )
            results.append(
                {
                    "branch_id": branch_id,
                    "row": upd["row"],
                    "status": "ok",
                    "meta_code": response.get("meta", {}).get("code"),
                    "from_hours": upd.get("branch_daily_hours"),
                    "to_hours": upd.get("hours_excel"),
                }
            )
        except Exception as exc:
            results.append(
                {
                    "branch_id": branch_id,
                    "row": upd["row"],
                    "status": "error",
                    "error": str(exc),
                    "from_hours": upd.get("branch_daily_hours"),
                    "to_hours": upd.get("hours_excel"),
                }
            )
    return results


def main():
    rows = read_excel_rows()
    yandex_failures = get_yandex_points(rows)

    branches = load_2gis_branches()
    branch_point_failures = enrich_branch_points(branches)

    matches, unmatched_rows, unmatched_branches = assign_by_distance(rows, branches, threshold_m=450)
    hour_updates = [m for m in matches if m["hour_mismatch"] and m.get("hours_excel")]

    report = {
        "excel_rows_total": len(rows),
        "excel_rows_with_yandex_point": sum(1 for r in rows if r.get("yandex_point")),
        "excel_rows_with_manual_or_url_point": sum(
            1
            for r in rows
            if r.get("yandex_point_source") in ("manual_columns", "url_params")
        ),
        "branches_total": len(branches),
        "branches_with_point": sum(1 for b in branches if b.get("point")),
        "matches": len(matches),
        "unmatched_rows": len(unmatched_rows),
        "unmatched_branches": len(unmatched_branches),
        "hour_mismatches": len(hour_updates),
        "yandex_point_failures": yandex_failures,
        "branch_point_failures": branch_point_failures,
        "match_details": matches,
        "unmatched_row_details": unmatched_rows,
        "unmatched_branch_details": unmatched_branches,
    }
    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    update_results = update_branch_hours(hour_updates)
    UPDATES_PATH.write_text(json.dumps(update_results, ensure_ascii=False, indent=2), encoding="utf-8")

    print(
        json.dumps(
            {
                "excel_rows_total": report["excel_rows_total"],
                "branches_total": report["branches_total"],
                "matches": report["matches"],
                "unmatched_rows": report["unmatched_rows"],
                "unmatched_branches": report["unmatched_branches"],
                "hour_mismatches": report["hour_mismatches"],
                "hour_updates_attempted": len(hour_updates),
                "hour_updates_ok": sum(1 for x in update_results if x["status"] == "ok"),
                "hour_updates_error": sum(1 for x in update_results if x["status"] != "ok"),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
