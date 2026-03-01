import json
from pathlib import Path
import urllib.parse
import urllib.request
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


TOKEN = "9ef36758fb840f45fd63ff5ca9ddf5c53feb4564"
ORG_ID = "70000001036827089"
CATALOG_KEY = "rupycp2722"
EXCEL_PATH = Path("data/stores_audit.xlsx")
DIFF_PATH = Path("data/2gis_sync_diff.json")


def api_get(url: str, headers=None):
    req = urllib.request.Request(url, headers=headers or {})
    with urllib.request.urlopen(req, timeout=40) as resp:
        return json.loads(resp.read().decode("utf-8"))


def daily_hours(schedule):
    if not isinstance(schedule, dict):
        return ""
    days = schedule.get("days")
    if not isinstance(days, dict):
        return ""
    vals = {(v.get("from"), v.get("to")) for v in days.values() if isinstance(v, dict)}
    if len(vals) != 1:
        return ""
    fr, to = next(iter(vals))
    if fr and to:
        return f"{fr}-{to}"
    return ""


def norm(text: str) -> str:
    return (text or "").lower().replace(" ", "").replace(",", "").replace(".", "")


def catalog_rating_votes(branch_id: str) -> str:
    if not str(branch_id).isdigit():
        return ""
    query = urllib.parse.urlencode(
        {
            "id": str(branch_id),
            "fields": "items.id,items.reviews",
            "key": CATALOG_KEY,
        }
    )
    url = f"https://catalog.api.2gis.com/3.0/items/byid?{query}"
    try:
        data = api_get(url)
        items = data.get("result", {}).get("items", [])
        if not items:
            return ""
        reviews = items[0].get("reviews") or {}
        rating = reviews.get("general_rating")
        votes = reviews.get("general_review_count_with_stars")
        if rating is None and votes is None:
            return ""
        if rating is not None and votes is not None:
            return f"{float(rating):.1f} ({int(votes)})"
        if rating is not None:
            return f"{float(rating):.1f}"
        return f"({int(votes)})"
    except Exception:
        return ""


def main():
    wb = load_workbook(EXCEL_PATH)
    source = wb[wb.sheetnames[0]]

    if "2gis" in wb.sheetnames:
        del wb["2gis"]
    ws = wb.create_sheet("2gis")

    headers = [
        "#",
        "Название в 2GIS",
        "Внутренний ID\n(заполнить)",
        "Адрес (2GIS)",
        "Ссылка 2GIS",
        "Время работы",
        "Дни работы",
        "Телефон",
        "2GIS Branch ID",
        "Статус / Разница",
        "Рейтинг 2GIS (оценка/голоса)",
        "Reviews Action (Real only)",
        "Review Note",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(1, i).value = h

    diff = json.loads(DIFF_PATH.read_text(encoding="utf-8"))
    matched = {int(m["row"]): str(m["branch_id"]) for m in diff.get("matches", []) if m.get("row") is not None}
    unmatched_by_row = {}
    for item in diff.get("unmatched_rows", []):
        row = item.get("row")
        if row is None:
            continue
        try:
            unmatched_by_row[int(row)] = item
        except Exception:
            continue

    # Manual overrides from latest safe additions and pending store additions.
    matched.update(
        {
            1: "branch_69a0979789091",
            7: "branch_69a09a71ee01d",
            10: "branch_69a09a72938e6",
            15: "branch_69a09a732b7ad",
            21: "branch_69a0976f94d5e",
            23: "branch_69a09a73bffbd",
            26: "branch_69a09a746a71b",
            30: "branch_69a08c05deaad",
        }
    )

    branches = api_get(
        f"https://api.account.2gis.com/api/1.0/branches?orgId={ORG_ID}&fields=schedule&pageSize=60&page=1",
        headers={"Authorization": f"Bearer {TOKEN}"},
    )["result"]["items"]
    branch_by_id = {str(b["id"]): b for b in branches}
    rating_by_id = {}
    for branch_id in {str(v) for v in matched.values() if str(v).isdigit()}:
        rating_by_id[branch_id] = catalog_rating_votes(branch_id)

    out_row = 2
    row_kind = {}

    def waiting_note(base_status: str, row_no: int) -> str:
        info = unmatched_by_row.get(row_no)
        if not info:
            return base_status

        reason = str(info.get("reason") or "").strip()
        if reason == "no_yandex_point":
            return f"{base_status}: нет pin в Яндексе (добавьте lat/lon или ll= в ссылке)"

        if reason == "distance_too_far_or_branch_already_used":
            nearest_id = str(info.get("nearest_branch_id") or "").strip() or "n/a"
            nearest_addr = str(info.get("nearest_branch_address") or "").strip()
            nearest_distance = info.get("nearest_distance_m")
            if isinstance(nearest_distance, (int, float)):
                dist_text = f"{int(round(float(nearest_distance)))}м"
            else:
                dist_text = "n/a"
            if nearest_addr:
                return f"{base_status}: ближайший 2GIS {nearest_id}, {dist_text}, {nearest_addr}"
            return f"{base_status}: ближайший 2GIS {nearest_id}, {dist_text}"

        if reason:
            return f"{base_status}: {reason}"
        return base_status
    for r in range(2, source.max_row + 1):
        row_no = source.cell(r, 1).value
        if not isinstance(row_no, (int, float)):
            continue
        row_no = int(row_no)

        brand_excel = source.cell(r, 2).value or ""
        internal_id = source.cell(r, 3).value or ""
        address_excel = source.cell(r, 4).value or ""
        hours_excel = source.cell(r, 6).value or ""

        branch_id = matched.get(row_no)
        branch = branch_by_id.get(branch_id) if branch_id else None

        name_2gis = branch.get("name", "") if branch else ""
        addr_2gis = branch.get("address", "") if branch else ""
        hours_2gis = daily_hours(branch.get("schedule")) if branch else ""
        days_2gis = "Ежедневно" if hours_2gis else ""
        link_2gis = f"https://2gis.uz/tashkent/firm/{branch_id}" if branch_id else ""

        if branch_id and str(branch_id).startswith("branch_"):
            status = "ожидаем модерацию"
            kind = "pending"
        elif branch_id:
            status = "есть в 2GIS"
            kind = "have"
        elif str(brand_excel).strip().upper() == "RUBA":
            status = "ждем информацию (RUBA)"
            kind = "ruba"
        else:
            status = "ждем информацию"
            kind = "waiting"

        diffs = []
        if branch:
            if hours_excel and hours_2gis and str(hours_excel).strip() != hours_2gis:
                diffs.append(f"Время: {hours_excel} != {hours_2gis}")
            if hours_excel and not hours_2gis:
                diffs.append("Время в 2GIS пусто")
            if str(brand_excel).strip() and "сырный домик" in str(brand_excel).lower() and "сырная" in name_2gis.lower():
                diffs.append("Бренд: в Excel Сырный Домик")
            if address_excel and addr_2gis:
                a = norm(str(address_excel))
                b = norm(str(addr_2gis))
                if a and b and a not in b and b not in a:
                    diffs.append("Адрес отличается")

        if status == "есть в 2GIS" and diffs:
            note = "есть в 2GIS, проверить: " + "; ".join(diffs)
            kind = "have_check"
        else:
            note = status
            if kind in ("waiting", "ruba"):
                note = waiting_note(note, row_no)

        if kind in ("have", "have_check", "pending"):
            reviews_action = "Request genuine customer reviews only; no incentives, no fake reviews"
            review_note = "Ask buyer to leave honest short feedback about real purchase/service"
        elif kind == "ruba":
            reviews_action = "RUBA track: complete listing ownership first"
            review_note = "After listing is active, start genuine review requests"
        else:
            reviews_action = "Create/confirm 2GIS listing first"
            review_note = "Reviews can start only after listing is live"

        ws.cell(out_row, 1).value = row_no
        ws.cell(out_row, 2).value = name_2gis
        ws.cell(out_row, 3).value = internal_id
        ws.cell(out_row, 4).value = addr_2gis
        ws.cell(out_row, 5).value = link_2gis
        ws.cell(out_row, 6).value = hours_2gis
        ws.cell(out_row, 7).value = days_2gis
        ws.cell(out_row, 8).value = "+998 78 555 15 15" if branch_id else ""
        ws.cell(out_row, 9).value = branch_id or ""
        ws.cell(out_row, 10).value = note
        ws.cell(out_row, 11).value = rating_by_id.get(str(branch_id), "") if branch_id else ""
        ws.cell(out_row, 12).value = reviews_action
        ws.cell(out_row, 13).value = review_note
        row_kind[out_row] = kind
        out_row += 1

    # Apply formatting similar to the first sheet.
    max_row = out_row - 1
    max_col = 13

    # Column widths A..M
    for col in "ABCDEFGHIJKLM":
        ws.column_dimensions[col].width = source.column_dimensions[col].width
    ws.column_dimensions["L"].width = 58
    ws.column_dimensions["M"].width = 56

    # Row heights and cell styles by row/column index.
    for r in range(1, max_row + 1):
        ws.row_dimensions[r].height = source.row_dimensions[r].height
        for c in range(1, max_col + 1):
            src_cell = source.cell(r, c)
            dst_cell = ws.cell(r, c)
            dst_cell._style = copy(src_cell._style)

    # Ensure J1..M1 headers are formatted like other headers.
    ws.cell(1, 10)._style = copy(source.cell(1, 1)._style)
    ws.cell(1, 11)._style = copy(source.cell(1, 1)._style)
    ws.cell(1, 12)._style = copy(source.cell(1, 1)._style)
    ws.cell(1, 13)._style = copy(source.cell(1, 1)._style)
    status_style = copy(source.cell(3, 10)._style)

    # Use explicit ARGB (FF...) to avoid transparent fills in Excel.
    def solid_fill(argb: str) -> PatternFill:
        return PatternFill(fill_type="solid", fgColor=argb)

    def color_font(argb: str) -> Font:
        return Font(color=argb)

    status_visuals = {
        "have": {
            "fill": solid_fill("FFC6EFCE"),
            "font": color_font("FF006100"),
        },
        "have_check": {
            "fill": solid_fill("FFFFF2CC"),
            "font": color_font("FF7F6000"),
        },
        "pending": {
            "fill": solid_fill("FFFCE4D6"),
            "font": color_font("FF9C5700"),
        },
        "waiting": {
            "fill": solid_fill("FFF8CBAD"),
            "font": color_font("FF9C0006"),
        },
        "ruba": {
            "fill": solid_fill("FFD9E1F2"),
            "font": color_font("FF1F4E78"),
        },
    }
    for r in range(2, max_row + 1):
        ws.cell(r, 10)._style = copy(status_style)
        visual = status_visuals.get(row_kind.get(r))
        if visual:
            ws.cell(r, 10).fill = copy(visual["fill"])
            ws.cell(r, 10).font = copy(visual["font"])

    # Keep links clickable in 2GIS column.
    for r in range(2, max_row + 1):
        link = ws.cell(r, 5).value
        if link:
            ws.cell(r, 5).hyperlink = str(link)

    ws.auto_filter.ref = f"A1:M{max_row}"
    ws.freeze_panes = "A2"

    wb.save(EXCEL_PATH)

    kind_counts = {}
    for v in row_kind.values():
        kind_counts[v] = kind_counts.get(v, 0) + 1

    stats = {
        "rows_total": out_row - 2,
        "mapped_total": sum(1 for r in range(2, out_row) if ws.cell(r, 9).value),
        "pending_total": sum(
            1 for r in range(2, out_row) if str(ws.cell(r, 9).value or "").startswith("branch_")
        ),
        "have_2gis": kind_counts.get("have", 0) + kind_counts.get("have_check", 0),
        "pending_moderation": kind_counts.get("pending", 0),
        "waiting_info": kind_counts.get("waiting", 0),
        "waiting_info_ruba": kind_counts.get("ruba", 0),
        "kind_counts": kind_counts,
    }
    print(json.dumps(stats, ensure_ascii=False))


if __name__ == "__main__":
    main()
